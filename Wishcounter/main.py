from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Color



"""Активация таблицы эксель"""
wb= Workbook()
ws=wb.active

"""Заголовки таблицы"""
ws["A1"]="Название События"
ws["B1"]="Время"
ws["C1"]="Ссылка на Событие"

"""Ширина столбцов"""
column_number = 1
column = str(chr(64 + column_number))
ws.column_dimensions[column].width = 60
column_number = 2
column = str(chr(64 + column_number))
ws.column_dimensions[column].width = 30
column_number = 3
column = str(chr(64 + column_number))
ws.column_dimensions[column].width = 20


request = requests.get("https://paimon.moe/timeline")
if request.status_code == 200:
    soup = BeautifulSoup(request.text,"html.parser")
    '''Добавление в таблицу данных со скачанной страницы'''
    i=1
    for text in soup.find_all(class_="flex items-center z-10 text-white cursor-pointer absolute rounded-l-xl rounded-r-xl"):
        i+=1
        if text.find(class_="text-sm rounded-xl text-black font-semibold bg-white bg-opacity-75 px-1"):
            """Изменение цвета GREEN"""
            c=ws["B" + str(i)]
            ft = Font(color="0000FF00")
            c.font=ft
            c = ws["A" + str(i)]
            c.font = ft
            ws["B"+str(i)]=(text.find(class_="text-sm rounded-xl text-black font-semibold bg-white bg-opacity-75 px-1")).get_text() # time
        else:
            """Изменение цвета RED"""
            c = ws["B" + str(i)]
            ft = Font(color="00FF0000")
            c.font = ft
            c = ws["A" + str(i)]
            c.font = ft
            ws["B"+str(i)]="Событие закончилось"
        ws["A"+str(i)]=text.find(class_="event-name text sticky left-0 font-display text-base md:text-lg text-black font-bold whitespace-nowrap overflow-hidden svelte-1wuqjd8").string # name


for text in soup.find_all(class_="text-primary hover:underline"):
    print(text)

"""Сохранение эксель таблицы"""
wb.save("WishCounter.xlsx")
print("File was successfully saved!")